import os
import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import pyexcel as p
from openpyxl import load_workbook
import csv
from mailService import send_email


addressTo: str = os.environ['MAIL_DESTINATION']
mailSubject: str = os.environ['MAIL_SUBJECT']
mailText: str = os.environ['MAIL_TEXT']
resultFile: str = os.environ['RESULT_FILE_NAME']

headers: dict = {'connection': 'keep-alive',
                 'cache-control': 'max-age=0',
                 'sec-ch-ua': '"Chromium";v="92", " Not A;Brand";v="99", "Google Chrome";v="92"',
                 'sec-ch-ua-mobile': '?0',
                 'origin': 'https://rmsp.nalog.ru',
                 'upgrade-insecure-requests': '1',
                 'DNT': '1',
                 'Content-Type': 'application/x-www-form-urlencoded',
                 'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36',
                 'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
                 'Sec-Fetch-Site': 'same-origin',
                 'Sec-Fetch-Mode': 'navigate',
                 'Sec-Fetch-User': '?1',
                 'Sec-Fetch-Dest': 'document',
                 'Referer': 'https://rmsp.nalog.ru/search.html?mode=inn-list',
                 'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
                 'Cookie': '_ym_uid=1626367445496233330; _ym_d=1626367445; _ym_isad=2; JSESSIONID=8E25F7E5601D4E5052DE1A5CC9763D4E', }


class OrganizationRecord:
    def __init__(self, number: int, fullName: str, innValue: int, ogrnValue: int):
        self.number = number
        self.fullName = fullName
        self.innValue = str(innValue)
        self.ogrnValue = str(ogrnValue)
        self.category = 'Значение'
        self.location = 'Значение'
        self.totalAmountOfTransfers = 'Значение'
        self.totalVolumeOfBankCommissions = 'Значение'
        self.amountSubsidies = 'Значение'

    def writeInfoAboutCategory(self, category: str) -> None:
        self.category = category

    def writeInfoAboutLocation(self, locationList: list) -> None:
        self.location = ' '.join(locationList)

    def returnOrganizationInfo(self) -> list:
        return [self.number,
                self.fullName,
                self.innValue,
                self.ogrnValue,
                self.category,
                self.location,
                self.totalAmountOfTransfers,
                self.totalVolumeOfBankCommissions,
                self.amountSubsidies]


def makeCsvFile() -> None:
    header: list = ['№ п/п',
                    'Полное наименование субъекта МСП',
                    'ИНН субъекта МСП',
                    'ОГРН субъекта МСП (при наличии)',
                    'Категория субъекта МСП (микро, малое, среднее)',
                    'Место нахождения (место жительства) субъекта МСП (субъект Российской Федерации)',
                    'Суммарный размер переводов, осуществленных физическими лицами в пользу субъектов МСП в СБП, рублей',
                    'Суммарный объём банковских комиссий за переводы денежных средств, осуществленных физическими лицами в '
                    'пользу субъектов МСП в СБП, рублей',
                    'Размер субсидий за отчётный период, рублей']
    with open(resultFile, 'w', encoding='UTF8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)


def makeRequest(innValue: int) -> requests.models.Response:
    session = requests.Session()
    retry = Retry(connect=3, backoff_factor=0.5)
    adapter = HTTPAdapter(max_retries=retry)
    session.mount('http://', adapter)
    session.mount('https://', adapter)
    response = session.get('https://rmsp.nalog.ru/report.xlsx?mode=inn-list&page=1&innList=' + str(
        innValue) + '&pageSize=10&sortField=NAME_EX&sort=ASC',
                           headers=headers)
    return response


def saveDataFromRequest(content: bytes) -> None:
    output = open('test.xlsx', 'wb')
    output.write(content)
    output.close()


def loadDataFromFile() -> list:
    workbook = load_workbook(filename='test.xlsx', read_only=False)
    worksheet = workbook.active
    data = []
    if worksheet is None:
        return data
    for curRow in worksheet.rows:
        data.append(list(map(lambda element: element.value, curRow)))
    workbook.close()
    return data


def findRowInData(innValue: str, ogrnValue: str, data: list) -> list:
    for row in data:
        if innValue in row and ogrnValue in row:
            return row
    return []


def writeRowIntoCsv(curOrganization) -> None:
    with open('result.csv', 'a', encoding='UTF8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(curOrganization.returnOrganizationInfo())


def main():
    makeCsvFile()
    p.save_book_as(file_name='SME+CORP+огрн.xls',
                   dest_file_name='SME+CORP+огрн.xlsx')
    workbook = load_workbook(filename='SME+CORP+огрн.xlsx', read_only=True)
    worksheet = workbook.active
    for row in worksheet.rows:
        if row[3].value == "C_INN":
            continue
        currentOrganization: object = OrganizationRecord(number=row[0].value,
                                                 fullName=row[2].value,
                                                 innValue=row[3].value,
                                                 ogrnValue=row[4].value)

        resultFromRequest: requests.models.Response = makeRequest(currentOrganization.innValue)
        if resultFromRequest.status_code != 200:
            print("Не было получено ответа от сайта ФНС - некорректный ИНН", currentOrganization.innValue)
            writeRowIntoCsv(currentOrganization)
            continue

        saveDataFromRequest(resultFromRequest.content)

        dataFromFile: list = loadDataFromFile()
        if not dataFromFile:
            print("По заданным параметрам не найдено сведений в едином реестре субъектов малого и среднего "
                  "предпринимательства. ИНН", currentOrganization.innValue)
            writeRowIntoCsv(currentOrganization)
            continue

        expectedRow: list = findRowInData(currentOrganization.innValue, currentOrganization.ogrnValue, dataFromFile)
        if not expectedRow:
            print("Не было найдено записи с ИНН", currentOrganization.innValue, "и ОГРН", currentOrganization.ogrnValue,
                  "в исходном документе")
            writeRowIntoCsv(currentOrganization)
            continue

        currentOrganization.writeInfoAboutCategory(expectedRow[3])
        currentOrganization.writeInfoAboutLocation(expectedRow[7:10])
        writeRowIntoCsv(currentOrganization)
    workbook.close()
    send_email(addressTo, mailSubject, mailText, resultFile)


if __name__ == "__main__":
    main()
